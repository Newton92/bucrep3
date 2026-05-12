# views_api_mailling_v3.py
import logging
import json
import os
from datetime import datetime
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils import timezone
from django.db import transaction
from rest_framework import generics, permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db.models import Q
from django.utils import timezone
from datetime import timedelta

from main.models import User, Commande, Acheteur, Document, MailInfo, MailAttachment, SuiviCommande
from main.serializers import (
    UserMailingSerializer, CommandeMailingSerializer, 
    DocumentSerializer, EnvoyerEmailSerializer
)
from rest_framework import generics, permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db.models import Q
from django.utils import timezone
from datetime import timedelta
from main.models import User, Commande, Acheteur, Document  # IMPORTANT: Ajouter Commande
from main.serializers import UserMailingSerializer, CommandeMailingSerializer, DocumentSerializer  # IMPORTANT: Ajouter CommandeMailingSerializer
from main.api.views_reporting import *  # Importer la fonction de génération de rapport
import re

from django.db.models import Count, Q
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
from collections import Counter

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from datetime import datetime
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from datetime import datetime
import os
import csv
import io
from django.http import HttpResponse



import logging

logger = logging.getLogger(__name__)

class ClientListAPIView(generics.ListAPIView):
    """
    API endpoint pour lister les clients actifs avec rôle Client
    URL: /api/gestion-des-mails/clients/
    Méthode: GET
    """
    serializer_class = UserMailingSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """
        Récupère la liste des utilisateurs:
        - avec le rôle 'Client'
        - actifs (is_active=True)
        - triés par nom d'utilisateur
        """
        try:
            logger.info(f"Recherche de clients avec rôle 'Client'")
            
            queryset = User.objects.filter(
                role__iexact='Client',  # Insensible à la casse
                is_active=True,
                activation=True
            ).order_by('username').distinct()
            
            logger.info(f"Nombre de clients trouvés (avant recherche): {queryset.count()}")
            
            # Recherche optionnelle par paramètre 'search'
            search_term = self.request.query_params.get('search', None)
            if search_term:
                logger.info(f"Recherche avec terme: {search_term}")
                queryset = queryset.filter(
                    Q(username__icontains=search_term) |
                    Q(first_name__icontains=search_term) |
                    Q(last_name__icontains=search_term) |
                    Q(email__icontains=search_term)
                )
                logger.info(f"Nombre de clients après recherche: {queryset.count()}")
            
            return queryset
            
        except Exception as e:
            logger.error(f"Erreur dans get_queryset: {str(e)}", exc_info=True)
            raise e
    
    def list(self, request, *args, **kwargs):
        """
        Personnalisation de la réponse
        """
        try:
            queryset = self.get_queryset()
            serializer = self.get_serializer(queryset, many=True)
            
            # Structure de réponse enrichie
            return Response({
                'status': 'success',
                'message': f'{len(serializer.data)} clients trouvés',
                'count': len(serializer.data),
                'data': serializer.data
            })
        except Exception as e:
            logger.error(f"Erreur dans list: {str(e)}", exc_info=True)
            return Response({
                'status': 'error',
                'message': f'Erreur serveur: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ClientCommandesAPIView(APIView):
    """
    API endpoint pour récupérer les commandes d'un client
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, client_id):
        try:
            logger.info(f"Récupération des commandes pour le client ID: {client_id}")
            
            # Récupérer le client
            try:
                client = User.objects.get(
                    id=client_id,
                    role__iexact='Client',
                    is_active=True
                )
                logger.info(f"Client trouvé: {client.username}")
            except User.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': 'Client non trouvé'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Récupérer les paramètres
            periode = request.query_params.get('periode', 'all')
            jours = request.query_params.get('jours', None)
            
            logger.info(f"Paramètres - Période: {periode}, Jours: {jours}")
            
            # Base queryset - SANS couleur_commentaire qui n'existe pas
            commandes = Commande.objects.filter(
                client=client,
                status__in=['nouvelle', 'en_cours']  # FILTRE PAR STATUT AJOUTÉ
            ).select_related(
                'acheteur',  # Garder seulement les relations qui existent
                'client'
            ).order_by('-date_recept_commande')
            
            logger.info(f"Commandes trouvées avant filtrage période: {commandes.count()}")
            
            # Filtrer par période
            commandes = self.filtrer_par_periode(commandes, periode, jours)
            
            logger.info(f"Commandes après filtrage période: {commandes.count()}")
            
            # Sérialiser les données
            serializer = CommandeMailingSerializer(commandes, many=True)
            
            # Statistiques
            stats = self.calculer_statistiques(commandes)
            
            return Response({
                'status': 'success',
                'message': f'{len(serializer.data)} commandes trouvées',
                'count': len(serializer.data),
                'stats': stats,
                'data': serializer.data
            })
            
        except User.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Client non trouvé'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Erreur dans ClientCommandesAPIView: {str(e)}", exc_info=True)
            return Response({
                'status': 'error',
                'message': f'Erreur serveur: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def filtrer_par_periode(self, queryset, periode, jours=None):
        """Filtre les commandes selon la période choisie"""
        aujourd_hui = timezone.now().date()
        
        try:
            if periode == 'today':
                # Aujourd'hui - Comparaison de dates sans __date
                return queryset.filter(
                    date_recept_commande__year=aujourd_hui.year,
                    date_recept_commande__month=aujourd_hui.month,
                    date_recept_commande__day=aujourd_hui.day
                )
            
            elif periode == '7days':
                # 7 derniers jours
                date_debut = aujourd_hui - timedelta(days=7)
                return queryset.filter(date_recept_commande__gte=date_debut)
            
            elif periode == '30days':
                # 30 derniers jours
                date_debut = aujourd_hui - timedelta(days=30)
                return queryset.filter(date_recept_commande__gte=date_debut)
            
            elif periode == 'month':
                # Ce mois
                return queryset.filter(
                    date_recept_commande__year=aujourd_hui.year,
                    date_recept_commande__month=aujourd_hui.month
                )
            
            elif periode == 'custom' and jours:
                # Période personnalisée
                try:
                    jours = int(jours)
                    date_debut = aujourd_hui - timedelta(days=jours)
                    return queryset.filter(date_recept_commande__gte=date_debut)
                except (ValueError, TypeError):
                    return queryset
            
            else:  # 'all' ou autre
                return queryset
                
        except Exception as e:
            logger.error(f"Erreur dans filtrer_par_periode: {str(e)}")
            return queryset
    
    def calculer_statistiques(self, queryset):
        """Calcule des statistiques sur les commandes"""
        try:
            total_commandes = queryset.count()
            
            if total_commandes == 0:
                return {
                    'total_commandes': 0,
                    'montant_total': 0,
                    'montant_total_formatted': '0 FCFA',
                    'moyenne_montant': 0,
                    'statuts': {}
                }
            
            # Montant total
            montant_total = 0
            statuts = {}
            
            for commande in queryset:
                # Additionner les montants
                if commande.credit_demande:
                    montant_total += float(commande.credit_demande)
                
                # Compter par statut
                statut = commande.status or 'non_defini'
                statuts[statut] = statuts.get(statut, 0) + 1
            
            # Formatage du montant
            montant_total_int = int(montant_total)
            montant_formatted = f"{montant_total_int:,} FCFA".replace(',', ' ')
            
            return {
                'total_commandes': total_commandes,
                'montant_total': montant_total_int,
                'montant_total_formatted': montant_formatted,
                'moyenne_montant': int(montant_total / total_commandes) if total_commandes > 0 else 0,
                'statuts': statuts
            }
            
        except Exception as e:
            logger.error(f"Erreur dans calculer_statistiques: {str(e)}")
            return {
                'total_commandes': 0,
                'montant_total': 0,
                'montant_total_formatted': '0 FCFA',
                'moyenne_montant': 0,
                'statuts': {}
            }
            
class AcheteurDocumentsAPIView(APIView):
    """
    API endpoint pour récupérer les documents d'un acheteur
    URL: /api/gestion-des-mails/acheteurs/<int:acheteur_id>/documents/
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, acheteur_id):
        try:
            logger.info(f"Récupération des documents pour l'acheteur ID: {acheteur_id}")
            
            # Récupérer l'acheteur
            try:
                acheteur = Acheteur.objects.get(id=acheteur_id)
                logger.info(f"Acheteur trouvé: {acheteur.nom}")
            except Acheteur.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': 'Acheteur non trouvé'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Récupérer les documents de l'acheteur
            documents = Document.objects.filter(
                acheteur=acheteur
            ).select_related(
                'created_by'  # Pour avoir l'utilisateur qui a uploadé
            ).order_by('-created_at')  # Changé de date_upload à created_at
            
            logger.info(f"Documents trouvés: {documents.count()}")
            
            # Sérialiser les documents
            serializer = DocumentSerializer(documents, many=True)
            
            # Statistiques sur les documents
            stats = self.calculer_statistiques(documents)
            
            return Response({
                'status': 'success',
                'message': f'{len(serializer.data)} documents trouvés',
                'count': len(serializer.data),
                'stats': stats,
                'acheteur': {
                    'id': acheteur.id,
                    'nom': acheteur.nom,
                    'code': acheteur.code
                },
                'data': serializer.data
            })
            
        except Exception as e:
            logger.error(f"Erreur dans AcheteurDocumentsAPIView: {str(e)}", exc_info=True)
            return Response({
                'status': 'error',
                'message': f'Erreur serveur: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def calculer_statistiques(self, queryset):
        """Calcule des statistiques sur les documents"""
        try:
            total_documents = queryset.count()
            
            if total_documents == 0:
                return {
                    'total_documents': 0,
                    'taille_totale': '0 Ko',
                    'types': {}
                }
            
            # Taille totale
            taille_totale = 0
            types = {}
            
            for doc in queryset:
                if doc.fichier and doc.fichier.size:
                    taille_totale += doc.fichier.size
                
                # Compter par extension
                if doc.fichier and doc.fichier.name:
                    extension = doc.fichier.name.split('.')[-1].lower() if '.' in doc.fichier.name else 'inconnu'
                    types[extension] = types.get(extension, 0) + 1
            
            # Formater la taille totale
            if taille_totale < 1024:
                taille_formatted = f"{taille_totale} o"
            elif taille_totale < 1024 * 1024:
                taille_formatted = f"{taille_totale / 1024:.1f} Ko"
            else:
                taille_formatted = f"{taille_totale / (1024 * 1024):.1f} Mo"
            
            return {
                'total_documents': total_documents,
                'taille_totale': taille_formatted,
                'types': types
            }
            
        except Exception as e:
            logger.error(f"Erreur dans calculer_statistiques documents: {str(e)}")
            return {
                'total_documents': 0,
                'taille_totale': '0 Ko',
                'types': {}
            }         
            
class GenererRapportEmailingAPIView(APIView):
    """
    API endpoint pour générer des rapports de solvabilité depuis l'emailing
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        try:
            data = request.data
            logger.info(f"Génération de rapport depuis l'emailing - Données reçues: {data}")
            
            commande_id = data.get('commande_id')
            format_rapport = data.get('format', 'pdf').lower()
            
            if not commande_id:
                return Response({
                    'status': 'error',
                    'message': 'ID de commande manquant'
                }, status=400)
            
            # Récupérer la commande
            try:
                commande = Commande.objects.get(id=commande_id)
                acheteur = commande.acheteur
            except Commande.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': 'Commande non trouvée'
                }, status=404)
            
            # TOUJOURS UTILISER LES ANNÉES PASSÉES (jamais l'année en cours)
            current_year = timezone.now().year
            years_to_use = [current_year - 1, current_year - 2, current_year - 3]
            
            logger.info(f"Années utilisées pour le rapport (toujours les années passées): {years_to_use}")
            
            # Récupérer la devise et le type de bilan
            devise_code = 'XAF'
            type_bilan = 'classique'
            
            try:
                from main.models import CompteFinancier
                compte_financier = CompteFinancier.objects.filter(acheteur=acheteur).first()
                
                if compte_financier:
                    if compte_financier.devise:
                        devise_code = compte_financier.devise
                    
                    if compte_financier.type_bilan:
                        type_bilan_map = {
                            'Classique': 'classique',
                            'classique': 'classique',
                            'Bancaire': 'bancaire',
                            'bancaire': 'bancaire',
                            'Anglais': 'anglais',
                            'anglais': 'anglais',
                            'Syscohada': 'syscohada',
                            'syscohada': 'syscohada',
                            'IFRS COBAC': 'ifrs',
                            'ifrs': 'ifrs',
                            'irfs_cobac': 'ifrs'
                        }
                        type_bilan = type_bilan_map.get(compte_financier.type_bilan, 'classique')
                        
            except Exception as e:
                logger.warning(f"Erreur récupération compte financier: {e}")
            
            # Préparer les données
            form_data = {
                'annee_n': years_to_use[0],
                'annee_n1': years_to_use[1],
                'annee_n2': years_to_use[2],
                'inclure_commande': 'oui',
                'commande_id': commande_id,
                'langue': data.get('langue', 'fr'),
                'devise': devise_code,
                'type_bilan': type_bilan,
                'format_rapport': format_rapport,
                'acheteur_id': acheteur.id
            }
            
            logger.info(f"Données préparées: {form_data}")
            
            # Appeler l'API de génération
            from django.test.client import RequestFactory
            import json
            
            factory = RequestFactory()
            host = request.get_host()
            
            fake_request = factory.post(
                '/api/reporting/generer-rapport-solvabilite/',
                data=json.dumps(form_data),
                content_type='application/json',
                HTTP_HOST=host,
                HTTP_AUTHORIZATION=request.META.get('HTTP_AUTHORIZATION', ''),
                HTTP_X_CSRFTOKEN=request.META.get('HTTP_X_CSRFTOKEN', '')
            )
            
            fake_request.user = request.user
            fake_request.session = request.session
            fake_request.COOKIES = request.COOKIES
            
            from main.api.views_reporting import generer_rapport_solvabilite
            response = generer_rapport_solvabilite(fake_request)
            
            if response.status_code != 200:
                logger.error(f"Erreur génération: {response.data}")
                return Response({
                    'status': 'error',
                    'message': 'Erreur lors de la génération du rapport'
                }, status=response.status_code)
            
            # Exporter
            from main.api.views_reporting import exporter_rapport_unifie
            
            report_data = response.data.get('report_data', {})
            
            export_response = exporter_rapport_unifie(
                request=request,
                report_data=report_data,
                form_data=form_data,
                export_format=format_rapport
            )
            
            return export_response
            
        except Exception as e:
            logger.error(f"Erreur: {str(e)}", exc_info=True)
            return Response({
                'status': 'error',
                'message': f'Erreur serveur: {str(e)}'
            }, status=500) 
  
class EnvoyerRapportEmailAPIView(APIView):
    """
    API endpoint pour envoyer un rapport de solvabilité par email
    URL: /api/emailing/envoyer/
    Méthode: POST
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
    def post(self, request):
        # Initialiser le logger avec l'utilisateur
        logger.info(f"📧 Tentative d'envoi d'email par l'utilisateur: {request.user.username}")
        
        # LOGS DÉTAILLÉS - AJOUTER CECI
        logger.info(f"🔍 Taille totale de la requête: {len(request.body)} bytes")
        logger.info(f"🔍 Content-Type: {request.content_type}")
        logger.info(f"🔍 FILES keys: {list(request.FILES.keys())}")
        
        # LOGS DÉTAILLÉS DES FICHIERS REÇUS
        logger.info(f"FILES reçus: {request.FILES.keys()}")
        for key, files in request.FILES.lists():
            for file in files:
                logger.info(f"Fichier reçu - Nom: {file.name}, Taille: {file.size}, Type: {file.content_type}")
        
        # LOGS DES DONNÉES
        logger.info(f"Data reçues: {request.data.keys()}")
        logger.info(f"commandes: {request.data.get('commandes')}")
        logger.info(f"documents: {request.data.get('documents')}")
        
        try:
            # =================================================================
            # 1. LOGS DÉTAILLÉS DES FICHIERS REÇUS
            # =================================================================
            logger.info(f"🔍 FILES reçus: {list(request.FILES.keys())}")
            
            fichiers_recus = []
            if 'fichiers' in request.FILES:
                fichiers_recus = request.FILES.getlist('fichiers')
                logger.info(f"✅ {len(fichiers_recus)} fichier(s) reçu(s) avec le champ 'fichiers'")
                for i, f in enumerate(fichiers_recus):
                    logger.info(f"   Fichier {i+1}: {f.name} - {f.size} bytes - {f.content_type}")
            else:
                logger.warning("⚠️ Aucun fichier reçu dans request.FILES['fichiers']")
            
            # =================================================================
            # 2. VALIDATION DES DONNÉES
            # =================================================================
            data = {
                'client_id': request.data.get('client_id'),
                'periode': request.data.get('periode'),
                'sujet': request.data.get('sujet'),
                'message': request.data.get('message'),
                'cc': request.data.get('cc', ''),
                'commandes': request.data.get('commandes', '[]'),
                'documents': request.data.get('documents', '[]'),
                'rapports': request.data.get('rapports', '[]'),
                'total_attachments': request.data.get('total_attachments', 0),
                'total_commands': request.data.get('total_commands', 0),
            }
            
            logger.info(f"📦 Données reçues: client_id={data['client_id']}, commandes={data['commandes']}")
            
            serializer = EnvoyerEmailSerializer(data=data)
            if not serializer.is_valid():
                logger.warning(f"❌ Données invalides: {serializer.errors}")
                return Response({
                    'status': 'error',
                    'message': 'Données invalides',
                    'errors': serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            validated_data = serializer.validated_data
            
            # =================================================================
            # 3. CRÉATION DU MAILINFO
            # =================================================================
            mail_info = MailInfo.objects.create(
                user=request.user,
                subject=validated_data['sujet'],
                cc_emails=validated_data.get('cc', ''),
                success=False
            )
            
            # Associer les commandes
            commandes = []
            if validated_data['commandes']:
                commandes = Commande.objects.filter(id__in=validated_data['commandes'])
                mail_info.commands.set(commandes)
                logger.info(f"📝 Commandes associées: {commandes.count()}")
            
            # =================================================================
            # 4. TRAITEMENT DES FICHIERS UPLOADÉS
            # =================================================================
            attachments = []
            
            # Récupérer les fichiers depuis request.FILES
            fichiers = request.FILES.getlist('fichiers')
            
            if fichiers:
                logger.info(f"📎 Traitement de {len(fichiers)} fichier(s) uploadé(s)...")
                
                for fichier in fichiers:
                    try:
                        # Vérifier la taille (max 10MB)
                        if fichier.size > 10 * 1024 * 1024:
                            logger.warning(f"⚠️ Fichier trop volumineux: {fichier.name} ({fichier.size} bytes)")
                            continue
                        
                        # Créer l'attachment
                        attachment = MailAttachment.objects.create(
                            upload=fichier,
                            mailinfo=mail_info
                        )
                        attachments.append(attachment)
                        logger.info(f"✅ Fichier uploadé attaché: {fichier.name} ({fichier.size} bytes)")
                        
                    except Exception as e:
                        logger.error(f"❌ Erreur lors de l'upload de {fichier.name}: {str(e)}")
            else:
                logger.info("ℹ️ Aucun fichier uploadé à traiter")
            
            # =================================================================
            # 5. TRAITEMENT DES DOCUMENTS ACHETEURS (par ID)
            # =================================================================
            if validated_data['documents']:
                logger.info(f"📄 Traitement de {len(validated_data['documents'])} document(s) par ID...")
                for doc_id in validated_data['documents']:
                    try:
                        doc = Document.objects.get(id=doc_id)
                        if doc.fichier and os.path.exists(doc.fichier.path):
                            # Créer un attachment à partir du document existant
                            attachment = MailAttachment.objects.create(
                                upload=doc.fichier,
                                mailinfo=mail_info
                            )
                            attachments.append(attachment)
                            logger.info(f"✅ Document attaché: {doc.titre} - {doc.fichier.name}")
                        else:
                            logger.warning(f"⚠️ Fichier du document {doc_id} non trouvé")
                    except Document.DoesNotExist:
                        logger.warning(f"⚠️ Document {doc_id} non trouvé")
            else:
                logger.info("ℹ️ Aucun document acheteur à traiter")
            
            # =================================================================
            # 6. ENVOI DE L'EMAIL AVEC PIÈCES JOINTES
            # =================================================================
            try:
                # Récupérer le client
                client = User.objects.get(id=validated_data['client_id'])
                
                # Créer l'email
                email = EmailMultiAlternatives(
                    subject=validated_data['sujet'],
                    body=re.sub(r'<[^>]+>', ' ', validated_data['message']),  # Version texte
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[client.email],
                    cc=validated_data.get('cc', '').split(';') if validated_data.get('cc') else [],
                )
                
                # Ajouter la version HTML
                email.attach_alternative(validated_data['message'], "text/html")
                
                # Ajouter les pièces jointes
                for attachment in attachments:
                    try:
                        # Récupérer le chemin du fichier
                        file_path = attachment.upload.path
                        
                        if os.path.exists(file_path):
                            with open(file_path, 'rb') as f:
                                file_content = f.read()
                                email.attach(
                                    filename=os.path.basename(attachment.upload.name),
                                    content=file_content,
                                    mimetype='application/octet-stream'
                                )
                            logger.info(f"✅ Pièce jointe ajoutée à l'email: {attachment.upload.name}")
                        else:
                            logger.warning(f"⚠️ Fichier non trouvé: {file_path}")
                            
                    except Exception as e:
                        logger.error(f"❌ Erreur lors de l'ajout de la pièce jointe: {str(e)}")
                
                # Envoyer l'email
                email.send(fail_silently=False)
                logger.info(f"✅ Email envoyé avec succès à {client.email}")
                
                # Mettre à jour le statut
                mail_info.success = True
                mail_info.save()
                
                # Mettre à jour les commandes
                for commande in commandes:
                    commande.status = 'envoye_client'
                    commande.date_envoi_client = timezone.now()
                    commande.email_envoye = True
                    commande.save()
                    
                    SuiviCommande.objects.create(
                        commande=commande,
                        user=request.user,
                        action=f"Rapport envoyé au client {client.username}",
                        type="ENVOI_CLIENT",
                        commentaire=f"Email envoyé avec {len(attachments)} pièce(s) jointe(s)"
                    )
                
                return Response({
                    'status': 'success',
                    'message': 'Email envoyé avec succès',
                    'mail_id': mail_info.id,
                    'pieces_jointes': len(attachments),
                    'details': {
                        'fichiers_uploades': len(fichiers) if fichiers else 0,
                        'documents_acheteurs': len(validated_data['documents']) if validated_data['documents'] else 0,
                        'total': len(attachments)
                    }
                }, status=status.HTTP_200_OK)
                
            except Exception as e:
                logger.error(f"❌ Erreur lors de l'envoi de l'email: {str(e)}", exc_info=True)
                mail_info.success = False
                mail_info.save()
                
                return Response({
                    'status': 'error',
                    'message': f'Erreur lors de l\'envoi: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Exception as e:
            logger.error(f"❌ Erreur inattendue: {str(e)}", exc_info=True)
            return Response({
                'status': 'error',
                'message': f'Erreur serveur: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HistoriqueEnvoisAPIView(APIView):
    """
    API endpoint pour récupérer l'historique des envois avec pagination
    URL: /api/emailing/historique/?page=1&page_size=10
    Méthode: GET
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        try:
            logger.info(f"Récupération de l'historique des envois pour {request.user.username}")
            
            # Paramètres de pagination
            page = int(request.query_params.get('page', 1))
            page_size = int(request.query_params.get('page_size', 10))
            
            # Valider les paramètres
            if page < 1:
                page = 1
            if page_size < 1 or page_size > 50:
                page_size = 10
            
            # Récupérer les envois
            queryset = MailInfo.objects.filter(
                user=request.user
            ).order_by('-date_sent')
            
            total_count = queryset.count()
            
            # Calculer les offsets
            start = (page - 1) * page_size
            end = start + page_size
            
            # Paginer
            envois = queryset[start:end]
            
            # Préparer les données
            data = []
            for envoi in envois:
                commandes = envoi.commands.all().values('id', 'reference_client', 'raison_sociale')
                pieces_jointes = envoi.mailattachment_set.count()
                
                data.append({
                    'id': envoi.id,
                    'date_sent': envoi.date_sent.strftime('%d/%m/%Y %H:%M'),
                    'subject': envoi.subject,
                    'success': envoi.success,
                    'commandes': list(commandes),
                    'pieces_jointes': pieces_jointes,
                    'cc_emails': envoi.get_cc_list()
                })
            
            # Calculer s'il y a une page suivante
            has_next = end < total_count
            
            logger.info(f"✅ {len(data)} envois trouvés (page {page}/{ (total_count + page_size - 1) // page_size })")
            
            return Response({
                'status': 'success',
                'count': len(data),
                'total': total_count,
                'page': page,
                'page_size': page_size,
                'has_next': has_next,
                'data': data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération de l'historique: {str(e)}", exc_info=True)
            return Response({
                'status': 'error',
                'message': f'Erreur serveur: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class DetailEnvoiAPIView(APIView):
    """
    API endpoint pour récupérer les détails d'un envoi spécifique
    URL: /api/emailing/historique/<int:envoi_id>/
    Méthode: GET
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, envoi_id):
        try:
            logger.info(f"Récupération des détails de l'envoi {envoi_id}")
            
            # Récupérer l'envoi
            envoi = MailInfo.objects.get(id=envoi_id, user=request.user)
            
            # Récupérer les commandes
            commandes = envoi.commands.all().values(
                'id', 'reference_client', 'raison_sociale', 'status', 'date_envoi_client'
            )
            
            # Récupérer les pièces jointes
            pieces_jointes = []
            for att in envoi.mailattachment_set.all():
                pieces_jointes.append({
                    'id': att.id,
                    'nom': os.path.basename(att.upload.name),
                    'taille': att.upload.size if att.upload else 0,
                    'url': att.upload.url if att.upload else None
                })
            
            data = {
                'id': envoi.id,
                'date_sent': envoi.date_sent.strftime('%d/%m/%Y %H:%M:%S'),
                'subject': envoi.subject,
                'success': envoi.success,
                'cc_emails': envoi.get_cc_list(),
                'commandes': list(commandes),
                'pieces_jointes': pieces_jointes,
                'formats_generes': envoi.formats_generes,
                'custom_days': envoi.custom_days
            }
            
            logger.info(f"✅ Détails de l'envoi {envoi_id} récupérés")
            
            return Response({
                'status': 'success',
                'data': data
            }, status=status.HTTP_200_OK)
            
        except MailInfo.DoesNotExist:
            logger.warning(f"Envoi {envoi_id} non trouvé")
            return Response({
                'status': 'error',
                'message': 'Envoi non trouvé'
            }, status=status.HTTP_404_NOT_FOUND)
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des détails: {str(e)}", exc_info=True)
            return Response({
                'status': 'error',
                'message': f'Erreur serveur: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
  

class ExporterHistoriqueAPIView(APIView):
    """
    API endpoint pour exporter l'historique des envois
    URL: /api/emailing/exporter-historique/?format=csv&type=all
    Format: csv ou excel
    Type: all, current_page, selected
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Remplacez les print par :
        logger.info(f"ExporterHistoriqueAPIView - User: {request.user}, Auth: {request.user.is_authenticated}")
        print(f"🔍 ExporterHistoriqueAPIView - User: {request.user}, Auth: {request.user.is_authenticated}")
        
        try:
            export_format = request.query_params.get('format', 'csv').lower()
            export_type   = request.query_params.get('type', 'all')
            selected_ids  = request.query_params.get('ids', '')
            page          = int(request.query_params.get('page', 1))
            page_size     = int(request.query_params.get('page_size', 10))

            queryset = MailInfo.objects.filter(
                user=request.user
            ).order_by('-date_sent')
            
            print(f"📊 Total envois trouvés: {queryset.count()}")

            # Filtrer selon le type
            if export_type == 'selected' and selected_ids:
                ids_list = [int(i) for i in selected_ids.split(',') if i.strip()]
                if ids_list:
                    queryset = queryset.filter(id__in=ids_list)
            elif export_type == 'current_page':
                start = (page - 1) * page_size
                end   = start + page_size
                queryset = queryset[start:end]

            # ✅ Même si queryset est vide, générer un fichier vide (pas d'erreur 404)
            rows = []
            for envoi in queryset:
                commandes_refs = ', '.join([
                    str(getattr(c, 'reference_client', None) or c.id)
                    for c in envoi.commandes.all()
                ])
                rows.append({
                    'ID'            : envoi.id,
                    'Date'          : envoi.date_sent.strftime('%d/%m/%Y %H:%M') if envoi.date_sent else '',
                    'Sujet'         : envoi.subject or '',
                    'Statut'        : 'Succès' if envoi.success else 'Échec',
                    'Commandes'     : commandes_refs,
                    'Pièces jointes': getattr(envoi, 'pieces_jointes', 0) or 0,
                })

            # ── CSV ──────────────────────────────────────────────────────────
            if export_format == 'csv':
                import csv
                from django.http import HttpResponse
                
                response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
                response['Content-Disposition'] = f'attachment; filename="historique_{export_type}.csv"'
                
                fieldnames = ['ID', 'Date', 'Sujet', 'Statut', 'Commandes', 'Pièces jointes']
                writer = csv.DictWriter(response, fieldnames=fieldnames, delimiter=';')
                writer.writeheader()
                writer.writerows(rows)
                
                print(f"✅ Export CSV généré: {len(rows)} lignes")
                return response

            # ── EXCEL ─────────────────────────────────────────────────────────
            elif export_format == 'excel':
                import io
                from django.http import HttpResponse
                
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                except ImportError:
                    return Response(
                        {'status': 'error', 'message': 'Installez openpyxl: pip install openpyxl'},
                        status=400
                    )

                wb = Workbook()
                ws = wb.active
                ws.title = "Historique"

                fieldnames = ['ID', 'Date', 'Sujet', 'Statut', 'Commandes', 'Pièces jointes']
                
                # En-têtes stylisés
                for col_idx, header in enumerate(fieldnames, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font      = Font(bold=True, color='FFFFFF')
                    cell.fill      = PatternFill('solid', fgColor='0D80BE')
                    cell.alignment = Alignment(horizontal='center')

                # Données
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, key in enumerate(fieldnames, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row.get(key, ''))

                # Largeur auto
                for col in ws.columns:
                    max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                response = HttpResponse(
                    buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="historique_{export_type}.xlsx"'
                print(f"✅ Export Excel généré: {len(rows)} lignes")
                return response

            else:
                return Response(
                    {'status': 'error', 'message': f'Format non supporté: {export_format}'},
                    status=400
                )

        except Exception as e:
            import traceback
            print(f"❌ Erreur export: {e}")
            print(traceback.format_exc())
            return Response({'status': 'error', 'message': str(e)}, status=500)
        
    
class StatistiquesAPIView(APIView):
    """
    API endpoint pour les statistiques
    URL: /api/emailing/statistiques/?period=30
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        try:
            period = int(request.query_params.get('period', 30))
            group_by = request.query_params.get('group_by', 'daily')
            
            # Date limite
            date_limit = timezone.now() - timedelta(days=period)
            
            # Récupérer les envois de la période
            envois = MailInfo.objects.filter(
                user=request.user,
                date_sent__gte=date_limit
            ).order_by('date_sent')
            
            # KPIs
            total = envois.count()
            success = envois.filter(success=True).count()
            failed = total - success
            success_rate = round((success / total * 100) if total > 0 else 0, 1)
            
            # Évolution temporelle
            if group_by == 'weekly':
                evolution = envois.annotate(periode=TruncWeek('date_sent'))
            elif group_by == 'monthly':
                evolution = envois.annotate(periode=TruncMonth('date_sent'))
            else:  # daily
                evolution = envois.annotate(periode=TruncDate('date_sent'))
            
            evolution_data = evolution.values('periode').annotate(
                total=Count('id'),
                reussis=Count('id', filter=Q(success=True)),
                echoues=Count('id', filter=Q(success=False))
            ).order_by('periode')
            
            # Distribution des commandes par envoi
            commandes_par_envoi = []
            pieces_par_envoi = []
            for envoi in envois:
                commandes_par_envoi.append(envoi.commands.count())
                pieces_par_envoi.append(envoi.mailattachment_set.count())
            
            # Top clients
            top_clients = Counter()
            for envoi in envois:
                for cmd in envoi.commands.all():
                    if cmd.client:
                        top_clients[cmd.client.username] += 1
            
            top_clients_data = dict(top_clients.most_common(5))
            
            return Response({
                'status': 'success',
                'kpis': {
                    'total': total,
                    'success': success,
                    'failed': failed,
                    'success_rate': success_rate
                },
                'evolution': {
                    'labels': [item['periode'].strftime('%d/%m/%Y') if group_by == 'daily' else 
                               item['periode'].strftime('Semaine %W %Y') if group_by == 'weekly' else
                               item['periode'].strftime('%B %Y') for item in evolution_data],
                    'reussis': [item['reussis'] for item in evolution_data],
                    'echoues': [item['echoues'] for item in evolution_data]
                },
                'distribution': {
                    'commandes': commandes_par_envoi,
                    'pieces': pieces_par_envoi
                },
                'top_clients': top_clients_data
            })
            
        except Exception as e:
            logger.error(f"Erreur stats: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=500)
        
                    

# Voir comment les commandes sont liées aux clients
class DiagnostiqueCommandesAPIView(APIView):
    """
    API endpoint pour diagnostiquer les relations entre commandes et clients
    URL: /api/gestion-des-mails/diagnostique-commandes/
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        try:
            # Voir la structure du modèle Commande
            commande_fields = [field.name for field in Commande._meta.get_fields()]
            
            # Voir quelques commandes
            commandes_sample = Commande.objects.all()[:5]
            commandes_data = []
            for cmd in commandes_sample:
                cmd_data = {
                    'id': cmd.id,
                    'numero_commande': cmd.numero_commande,
                    'date_commande': str(cmd.date_commande) if cmd.date_commande else None,
                }
                
                # Voir les relations possibles
                if hasattr(cmd, 'client') and cmd.client:
                    cmd_data['client_id'] = cmd.client.id
                    cmd_data['client_username'] = cmd.client.username
                
                if hasattr(cmd, 'user') and cmd.user:
                    cmd_data['user_id'] = cmd.user.id
                    cmd_data['user_username'] = cmd.user.username
                
                if hasattr(cmd, 'acheteur') and cmd.acheteur:
                    cmd_data['acheteur_id'] = cmd.acheteur.id
                    cmd_data['acheteur_nom'] = cmd.acheteur.nom
                    
                    # Voir si l'acheteur a un client
                    if hasattr(cmd.acheteur, 'client') and cmd.acheteur.client:
                        cmd_data['acheteur_client_id'] = cmd.acheteur.client.id
                
                commandes_data.append(cmd_data)
            
            return Response({
                'status': 'success',
                'champs_commande': commande_fields,
                'total_commandes': Commande.objects.count(),
                'commandes_sample': commandes_data
            })
            
        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=500)

# Vue de test
class TestMailingAPIView(APIView):
    """
    API endpoint de test
    URL: /api/gestion-des-mails/test/
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        return Response({
            'status': 'success',
            'message': 'L\'API de gestion des emails est opérationnelle',
            'user': request.user.username
        })