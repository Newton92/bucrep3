"""API views pour le module Gestion des Exportations (Listing & Décompte)."""
import csv
import io
from datetime import datetime

from django.http import HttpResponse
from django.db.models import Q, Prefetch
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from main.models import (
    Acheteur, AdresseAcheteur, PortableAcheteur, EmailAcheteur,
    Resume, Commande,
)


# ─────────────────────────────────────────────────────────────
#  Helpers Excel
# ─────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="0D80BE")
_HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN   = Alignment(vertical="center", wrap_text=True)
_THIN         = Side(style="thin", color="D0D7DE")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_excel(ws, headers, rows, col_widths=None):
    """Remplit une feuille Excel avec headers stylisés + données."""
    ws.row_dimensions[1].height = 26

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = (col_widths or {}).get(col_idx, 22)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _DATA_ALIGN
            cell.border = _BORDER

    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────
#  Helpers données
# ─────────────────────────────────────────────────────────────

def _get_listing_queryset(params):
    qs = Acheteur.objects.select_related("pays", "ville", "province") \
        .prefetch_related(
            Prefetch("adresses",  queryset=AdresseAcheteur.objects.all()),
            Prefetch("portables", queryset=PortableAcheteur.objects.all()),
            Prefetch("resume_set", queryset=Resume.objects.order_by("-created_at")),
        )

    domaine    = params.get("domaine", "").strip()
    search     = params.get("search", "").strip()

    if domaine:
        qs = qs.filter(activite_principale__icontains=domaine)
    if search:
        qs = qs.filter(Q(nom__icontains=search) | Q(code__icontains=search))

    return qs.order_by("nom")


def _listing_row(acheteur):
    adresse  = acheteur.adresses.first()
    portable = acheteur.portables.first()
    resume   = acheteur.resume_set.first()
    return [
        acheteur.nom or "",
        adresse.adresse if adresse else (
            " ".join(filter(None, [acheteur.numero_adresse, acheteur.rue_adresse])) or ""
        ),
        portable.portable if portable else "",
        acheteur.email or "",
        acheteur.code or "",
        int(resume.chiffre_affaire) if (resume and resume.chiffre_affaire) else "",
    ]


def _get_decompte_queryset(params):
    qs = Commande.objects.select_related("acheteur", "client", "pays")

    acheteur   = params.get("acheteur", "").strip()
    sigle      = params.get("sigle", "").strip()
    client_id  = params.get("client_id", "").strip()
    date_debut = params.get("date_debut", "").strip()
    date_fin   = params.get("date_fin", "").strip()

    if acheteur:
        qs = qs.filter(acheteur__nom__icontains=acheteur)
    if sigle:
        qs = qs.filter(acheteur__sigle__icontains=sigle)
    if client_id:
        qs = qs.filter(client_id=client_id)
    if date_debut:
        qs = qs.filter(date_recept_commande__gte=date_debut)
    if date_fin:
        qs = qs.filter(date_recept_commande__lte=date_fin)

    return qs.order_by("-date_recept_commande")


def _decompte_row(cmd):
    return [
        cmd.acheteur.nom if cmd.acheteur else "",
        cmd.client_nom or (cmd.client.get_full_name() if cmd.client else ""),
        cmd.date_recept_commande.strftime("%d/%m/%Y") if cmd.date_recept_commande else "",
        cmd.date_rapport.strftime("%d/%m/%Y")         if cmd.date_rapport          else "",
        cmd.reference_client or "",
        cmd.notre_ref or "",
        cmd.type_rapport or "",
        cmd.pays.nom if cmd.pays else "",
    ]


# ─────────────────────────────────────────────────────────────
#  API — Listing
# ─────────────────────────────────────────────────────────────

LISTING_HEADERS = ["BUYER", "ADRESSE", "PORTABLE", "COURRIEL", "N° FISCAL", "CA"]
LISTING_WIDTHS  = {1: 30, 2: 40, 3: 18, 4: 30, 5: 18, 6: 18}


class ExportListingDataAPIView(APIView):
    """GET /api/exports/listing/ — données JSON paginées."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs   = _get_listing_queryset(request.query_params)
        page = int(request.query_params.get("page", 1))
        per  = int(request.query_params.get("per_page", 50))
        total = qs.count()

        start = (page - 1) * per
        end   = start + per
        rows  = [_listing_row(a) for a in qs[start:end]]

        return Response({
            "total":    total,
            "page":     page,
            "per_page": per,
            "pages":    (total + per - 1) // per,
            "headers":  LISTING_HEADERS,
            "rows":     rows,
        })


class ExportListingFileAPIView(APIView):
    """GET /api/exports/listing/export/?format=excel|csv — téléchargement fichier."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fmt = request.query_params.get("format", "excel").lower()
        qs  = _get_listing_queryset(request.query_params)
        rows = [_listing_row(a) for a in qs]
        ts   = datetime.now().strftime("%Y%m%d_%H%M")

        if fmt == "csv":
            response = HttpResponse(content_type="text/csv; charset=utf-8")
            response["Content-Disposition"] = f'attachment; filename="listing_acheteurs_{ts}.csv"'
            response.write("﻿")  # BOM UTF-8
            writer = csv.writer(response, delimiter=";")
            writer.writerow(LISTING_HEADERS)
            writer.writerows(rows)
            return response

        # Excel (défaut)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listing Acheteurs"
        _write_excel(ws, LISTING_HEADERS, rows, LISTING_WIDTHS)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="listing_acheteurs_{ts}.xlsx"'
        return response


# ─────────────────────────────────────────────────────────────
#  API — Décompte
# ─────────────────────────────────────────────────────────────

DECOMPTE_HEADERS = [
    "BUYER", "CUSTOMER", "DATE RECEIVED", "REPORT DATE",
    "CLIENT REFERENCE", "OUR REFERENCE", "REPORT TYPE", "COUNTRY",
]
DECOMPTE_WIDTHS = {1: 30, 2: 28, 3: 16, 4: 16, 5: 22, 6: 22, 7: 22, 8: 20}


class ExportDecompteDataAPIView(APIView):
    """GET /api/exports/decompte/ — données JSON paginées."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs    = _get_decompte_queryset(request.query_params)
        page  = int(request.query_params.get("page", 1))
        per   = int(request.query_params.get("per_page", 50))
        total = qs.count()

        start = (page - 1) * per
        end   = start + per
        rows  = [_decompte_row(c) for c in qs[start:end]]

        return Response({
            "total":    total,
            "page":     page,
            "per_page": per,
            "pages":    (total + per - 1) // per,
            "headers":  DECOMPTE_HEADERS,
            "rows":     rows,
        })


class ExportDecompteFileAPIView(APIView):
    """GET /api/exports/decompte/export/?format=excel|csv — téléchargement fichier."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fmt  = request.query_params.get("format", "excel").lower()
        qs   = _get_decompte_queryset(request.query_params)
        rows = [_decompte_row(c) for c in qs]
        ts   = datetime.now().strftime("%Y%m%d_%H%M")

        if fmt == "csv":
            response = HttpResponse(content_type="text/csv; charset=utf-8")
            response["Content-Disposition"] = f'attachment; filename="decompte_commandes_{ts}.csv"'
            response.write("﻿")
            writer = csv.writer(response, delimiter=";")
            writer.writerow(DECOMPTE_HEADERS)
            writer.writerows(rows)
            return response

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Décompte Commandes"
        _write_excel(ws, DECOMPTE_HEADERS, rows, DECOMPTE_WIDTHS)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="decompte_commandes_{ts}.xlsx"'
        return response
